import json
import sys
import re
import os
import tempfile
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


# Expected field names in mapping objects from AI response.
EXPECTED_MAPPING_FIELDS = [
    "source_wb_keyword",
    "source_ws",
    "item_col_position",
    "item_col_name",
    "forecast_col_position",
    "forecast_col_name",
    "data_start_row",
]

NUMERIC_MAPPING_FIELDS = {
    "item_col_position",
    "forecast_col_position",
    "data_start_row",
}


def _extract_response_text(response_text: str) -> str:
    """
    Unwrap response text when a full record JSON is passed instead of raw LLM text.
    """
    text = response_text.strip()
    if not text:
        return ""

    try:
        payload = json.loads(text)
    except (json.JSONDecodeError, ValueError, TypeError):
        return response_text

    if isinstance(payload, dict):
        llm_response = payload.get("llm_response")
        if isinstance(llm_response, dict):
            nested = llm_response.get("response")
            if isinstance(nested, str) and nested.strip():
                return nested
        elif isinstance(llm_response, str) and llm_response.strip():
            return llm_response

        for key in ("response", "output", "content"):
            candidate = payload.get(key)
            if isinstance(candidate, str) and candidate.strip():
                return candidate

    return response_text


def _normalize_mapping_fields(mapping: dict[str, Any]) -> dict[str, Any]:
    """
    Normalize mapping keys to expected snake_case field names.
    """
    normalized: dict[str, Any] = {}

    for key, value in mapping.items():
        cleaned = key.strip().lower().replace(" ", "_")
        if cleaned == "skipped":
            normalized["skipped"] = value
        elif cleaned in EXPECTED_MAPPING_FIELDS:
            normalized[cleaned] = value

    return normalized


def _sheet_has_expected_headers(ws) -> bool:
    """ Function checks that the expected input headers are the same after normalization
    """
    headers = [
        ws.cell(row=1, column=index + 1).value
        for index in range(len(EXPECTED_MAPPING_FIELDS))
    ]
    normalized_headers = [
        str(value).strip().lower() if value is not None else ""
        for value in headers
    ]
    return normalized_headers == EXPECTED_MAPPING_FIELDS


def _pick_target_sheet_name(wb) -> str | None:
    """
    Choose a config sheet for row appends.
    """
    preferred = "utility_configs"
    if preferred in wb.sheetnames and _sheet_has_expected_headers(wb[preferred]):
        return preferred

    for sheet_name in wb.sheetnames:
        if _sheet_has_expected_headers(wb[sheet_name]):
            return sheet_name

    if wb.sheetnames:
        return wb.sheetnames[0]
    return None


def _safe_cell_value(value: Any) -> Any:
    """ Safety function that converts values to JSON strings if the value is a 'list' or 'dict'
    """
    if isinstance(value, (list, dict)):
        return json.dumps(value, ensure_ascii=False)
    return value


def _normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip().lower()


def _coerce_mapping_value(field: str, value: Any) -> Any:
    if value is None:
        return None

    if isinstance(value, str):
        trimmed = value.strip()
        if not trimmed or trimmed.lower() == "none":
            return None
        value = trimmed

    if field in NUMERIC_MAPPING_FIELDS and value is not None:
        try:
            return int(value)
        except (TypeError, ValueError):
            return value

    return value


def _coerce_mapping_row(mapping: dict[str, Any]) -> dict[str, Any]:
    coerced: dict[str, Any] = {}
    for field in EXPECTED_MAPPING_FIELDS:
        coerced[field] = _coerce_mapping_value(field, mapping.get(field))
    return coerced


def _row_signature(row: dict[str, Any]) -> str:
    parts = [_normalize_text(row.get(field)) for field in EXPECTED_MAPPING_FIELDS]
    return "|".join(parts)


def _row_similarity(left: dict[str, Any], right: dict[str, Any]) -> float:
    return SequenceMatcher(None, _row_signature(left), _row_signature(right)).ratio()


def _load_existing_rows(wb, eligible_sheets: list[str]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for sheet_name in eligible_sheets:
        ws = wb[sheet_name]
        for row_index in range(2, ws.max_row + 1):
            raw = {
                field: ws.cell(row=row_index, column=idx + 1).value
                for idx, field in enumerate(EXPECTED_MAPPING_FIELDS)
            }
            coerced = _coerce_mapping_row(raw)
            if all(value is None for value in coerced.values()):
                continue
            rows.append(
                {
                    "sheet_name": sheet_name,
                    "row_index": row_index,
                    "row": coerced,
                }
            )
    return rows


def _eligible_config_sheets(wb) -> list[str]:
    return [sheet_name for sheet_name in wb.sheetnames if _sheet_has_expected_headers(wb[sheet_name])]


def _extract_config_from_record_payload(response_text: str) -> tuple[str | None, str]:
    """
    If full record JSON text is provided, return (config_file_path, llm_response_text).
    Otherwise returns (None, unwrapped_response_text).
    """
    text = response_text.strip()
    if not text:
        return None, ""

    try:
        payload = json.loads(text)
    except (json.JSONDecodeError, ValueError, TypeError):
        return None, _extract_response_text(response_text)

    if not isinstance(payload, dict):
        return None, _extract_response_text(response_text)

    input_data = payload.get("input") if isinstance(payload.get("input"), dict) else {}
    config_file_path = (
        payload.get("config_file_path")
        or payload.get("config_file_name")
        or input_data.get("config_file_path")
        or input_data.get("config_file_name")
    )

    return config_file_path, _extract_response_text(response_text)


def build_smart_update_preview(config_file_path: str, response_text: str, fuzzy_threshold: float = 0.85) -> dict[str, Any]:
    unwrapped_response = _extract_response_text(response_text)
    raw_mappings = _extract_raw_mappings(unwrapped_response)
    if not raw_mappings:
        return {
            "success": False,
            "error": "Malformed or unsupported LLM response; no mappings could be parsed.",
            "proposals": [],
        }

    config_path = _resolve_config_path(config_file_path)
    if not config_path.exists():
        return {
            "success": False,
            "error": f"Config file not found: {config_path}",
            "proposals": [],
        }

    if config_path.suffix.lower() not in {".xlsx", ".xls", ".xlsm"}:
        return {
            "success": False,
            "error": f"Unsupported file type: {config_path.suffix}",
            "proposals": [],
        }

    try:
        wb = load_workbook(config_path)
    except Exception as error:
        return {
            "success": False,
            "error": f"Failed to load workbook: {error}",
            "proposals": [],
        }

    eligible_sheets = _eligible_config_sheets(wb)
    if not eligible_sheets:
        return {
            "success": False,
            "error": "No config worksheet has expected headers for smart update.",
            "proposals": [],
        }

    existing_rows = _load_existing_rows(wb, eligible_sheets)
    proposals: list[dict[str, Any]] = []

    for index, raw_mapping in enumerate(raw_mappings, start=1):
        if not isinstance(raw_mapping, dict):
            continue

        mapping = _normalize_mapping_fields(raw_mapping)

        mapping_keys = {str(key).strip().lower().replace(" ", "_") for key in raw_mapping.keys()}
        missing_fields = [field for field in EXPECTED_MAPPING_FIELDS if field not in mapping_keys]
        extra_fields = [
            key for key in raw_mapping.keys()
            if str(key).strip().lower().replace(" ", "_") not in EXPECTED_MAPPING_FIELDS
            and str(key).strip().lower().replace(" ", "_") != "skipped"
        ]

        proposed_row = _coerce_mapping_row(mapping)
        change_id = f"change_{index}"

        if missing_fields or extra_fields:
            errors = []
            if missing_fields:
                errors.append(f"Missing required fields: {', '.join(missing_fields)}")
            if extra_fields:
                errors.append(f"Unexpected fields: {', '.join(sorted(extra_fields))}")
            proposals.append(
                {
                    "change_id": change_id,
                    "status": "rejected",
                    "action": "rejected",
                    "reason": "row_schema_mismatch",
                    "errors": errors,
                    "before_row": None,
                    "after_row": proposed_row,
                    "target_sheet": None,
                    "target_row_index": None,
                    "similarity": None,
                    "available_target_sheets": eligible_sheets,
                }
            )
            continue

        exact_match = None
        for existing in existing_rows:
            if _row_signature(existing["row"]) == _row_signature(proposed_row):
                exact_match = existing
                break

        if exact_match:
            proposals.append(
                {
                    "change_id": change_id,
                    "status": "proposed",
                    "action": "update_exact",
                    "reason": "exact_full_row_match",
                    "errors": [],
                    "before_row": exact_match["row"],
                    "after_row": proposed_row,
                    "target_sheet": exact_match["sheet_name"],
                    "target_row_index": exact_match["row_index"],
                    "similarity": 1.0,
                    "available_target_sheets": [exact_match["sheet_name"]],
                }
            )
            continue

        default_sheet = eligible_sheets[0] if len(eligible_sheets) == 1 else None
        proposals.append(
            {
                "change_id": change_id,
                "status": "proposed",
                "action": "append_new",
                "reason": "no_match_found",
                "errors": [],
                "before_row": None,
                "after_row": proposed_row,
                "target_sheet": default_sheet,
                "target_row_index": None,
                "similarity": None,
                "available_target_sheets": eligible_sheets,
            }
        )

    return {
        "success": True,
        "config_file": str(config_path),
        "fuzzy_threshold": fuzzy_threshold,
        "eligible_sheets": eligible_sheets,
        "proposals": proposals,
        "summary": {
            "total": len(proposals),
            "exact": len([p for p in proposals if p.get("action") == "update_exact"]),
            "fuzzy": 0,
            "append": len([p for p in proposals if p.get("action") == "append_new"]),
            "rejected": len([p for p in proposals if p.get("action") == "rejected"]),
        },
    }


def apply_smart_update_changes(
    config_file_path: str,
    response_text: str,
    accepted_changes: list[dict[str, Any] | str],
    fuzzy_threshold: float = 0.85,
) -> dict[str, Any]:
    preview = build_smart_update_preview(config_file_path, response_text, fuzzy_threshold=fuzzy_threshold)
    if not preview.get("success"):
        return preview

    proposals = preview.get("proposals") if isinstance(preview.get("proposals"), list) else []
    eligible_sheets = preview.get("eligible_sheets") if isinstance(preview.get("eligible_sheets"), list) else []

    accepted_map: dict[str, dict[str, Any]] = {}
    for entry in accepted_changes:
        if isinstance(entry, str):
            accepted_map[entry] = {}
        elif isinstance(entry, dict):
            change_id = str(entry.get("change_id") or "").strip()
            if change_id:
                accepted_map[change_id] = entry

    config_path = Path(preview.get("config_file") or "")
    try:
        wb = load_workbook(config_path)
    except Exception as error:
        return {
            "success": False,
            "error": f"Failed to load workbook for apply: {error}",
        }

    updated = 0
    added = 0
    skipped = 0
    rejected = 0
    applied_change_ids: list[str] = []
    skipped_details: list[dict[str, Any]] = []

    for proposal in proposals:
        change_id = str(proposal.get("change_id") or "")
        action = proposal.get("action")
        if not change_id:
            continue

        if action == "rejected":
            rejected += 1

        accepted_payload = accepted_map.get(change_id)
        if accepted_payload is None:
            skipped += 1
            skipped_details.append({"change_id": change_id, "reason": "not_accepted"})
            continue

        if action == "rejected":
            return {
                "success": False,
                "error": f"Rejected proposal {change_id} cannot be applied.",
            }

        after_row = proposal.get("after_row") if isinstance(proposal.get("after_row"), dict) else {}
        row_values = [_safe_cell_value(after_row.get(field)) for field in EXPECTED_MAPPING_FIELDS]

        if action == "update_exact":
            target_sheet = proposal.get("target_sheet")
            target_row_index = proposal.get("target_row_index")
            if target_sheet not in wb.sheetnames:
                return {
                    "success": False,
                    "error": f"Target worksheet not found for {change_id}: {target_sheet}",
                }
            if not isinstance(target_row_index, int) or target_row_index < 2:
                return {
                    "success": False,
                    "error": f"Invalid target row for {change_id}: {target_row_index}",
                }

            ws = wb[target_sheet]
            for column_index, value in enumerate(row_values, start=1):
                ws.cell(row=target_row_index, column=column_index, value=value)
            updated += 1
            applied_change_ids.append(change_id)
            continue

        if action == "append_new":
            selected_sheet = accepted_payload.get("target_sheet")
            if not selected_sheet:
                selected_sheet = proposal.get("target_sheet")

            if selected_sheet not in eligible_sheets:
                return {
                    "success": False,
                    "error": (
                        f"A valid worksheet selection is required for {change_id}. "
                        f"Valid sheets: {', '.join(eligible_sheets)}"
                    ),
                }

            ws = wb[selected_sheet]
            ws.append(row_values)
            added += 1
            applied_change_ids.append(change_id)
            continue

        skipped += 1
        skipped_details.append({"change_id": change_id, "reason": f"unsupported_action:{action}"})

    temp_file = None
    try:
        with tempfile.NamedTemporaryFile(
            delete=False,
            suffix=config_path.suffix,
            prefix=f"{config_path.stem}_smart_update_",
            dir=str(config_path.parent),
        ) as handle:
            temp_file = Path(handle.name)

        wb.save(temp_file)
        os.replace(temp_file, config_path)
    except Exception as error:
        if temp_file and temp_file.exists():
            temp_file.unlink(missing_ok=True)
        return {
            "success": False,
            "error": f"Failed to save config updates: {error}",
            "applied_change_ids": applied_change_ids,
        }

    return {
        "success": True,
        "config_file": str(config_path),
        "updated_count": updated,
        "added_count": added,
        "skipped_count": skipped,
        "rejected_count": rejected,
        "applied_change_ids": applied_change_ids,
        "skipped_details": skipped_details,
    }


def _resolve_config_path(config_file: str) -> Path:
    """
    Resolve config file path with Docker and local fallbacks.
    Tries:
      1. Absolute path if provided
      2. /data/config/ (Docker mounted path)
      3. <repo_root>/data/config/ (local path)
    """
    target = Path(config_file)

    # If absolute path exists, use it.
    if target.is_absolute() and target.exists():
        return target

    # Try Docker mounted path.
    docker_path = Path("/data/config") / target.name
    if docker_path.exists():
        return docker_path

    # Try local repo path (assume we're in backend/src, go up to root, then data/config).
    current_dir = Path(__file__).resolve().parent
    repo_root = current_dir.parent.parent
    local_path = repo_root / "data" / "config" / target.name
    if local_path.exists():
        return local_path

    # If none exist, return first attempted path for error handling.
    if target.is_absolute():
        return target
    return local_path


def _extract_json_mappings(response_text: str) -> list[dict[str, Any]]:
    """
    Attempt to extract and parse JSON mappings from response text.
    Returns list of mapping dictionaries if successful, else empty list.
    """
    try:
        # Try direct JSON parse.
        data = json.loads(response_text)
        if isinstance(data, dict) and "mappings" in data:
            return data["mappings"] if isinstance(data["mappings"], list) else []
        if isinstance(data, list):
            return data
        return []
    except (json.JSONDecodeError, ValueError):
        return []


def _extract_legacy_format_mappings(response_text: str) -> list[dict[str, Any]]:
    """
    Fall back to legacy regex-based extraction for OUTPUT blocks.
    Handles format: "OUTPUT : <key>: <value>" lines.
    """
    mappings = []
    # Split by OUTPUT blocks.
    output_blocks = re.split(r'OUTPUT\s*:', response_text, flags=re.IGNORECASE)

    for block in output_blocks[1:]:  # Skip first element (before first OUTPUT).
        mapping = {}
        # Extract key: value pairs.
        for line in block.split("\n"):
            line = line.strip()
            if not line:
                continue
            if ":" in line:
                key, value = line.split(":", 1)
                mapping[key.strip()] = value.strip()

        # Only add if it has content.
        if mapping:
            mappings.append(mapping)

    return mappings


def _extract_plain_key_value_mappings(response_text: str) -> list[dict[str, Any]]:
    """
    Extract mappings from plain key:value text blocks separated by blank lines.
    Supports wrapped values where continuation lines do not include a colon.
    """
    mappings: list[dict[str, Any]] = []
    current: dict[str, Any] = {}
    current_key: str | None = None

    def _flush_current() -> None:
        nonlocal current, current_key
        if current:
            mappings.append(current)
        current = {}
        current_key = None

    for raw_line in response_text.splitlines():
        line = raw_line.strip()
        if not line:
            _flush_current()
            continue

        if ":" in line:
            key_part, value_part = line.split(":", 1)
            normalized_key = key_part.strip().lower().replace(" ", "_")
            if normalized_key in EXPECTED_MAPPING_FIELDS or normalized_key == "skipped":
                if normalized_key in current and current:
                    _flush_current()
                current[normalized_key] = value_part.strip()
                current_key = normalized_key
                continue

        if current_key and line:
            previous = str(current.get(current_key, "")).strip()
            current[current_key] = f"{previous} {line}".strip()

    _flush_current()
    return mappings


def _extract_raw_mappings(unwrapped_response: str) -> list[dict[str, Any]]:
    mappings = _extract_json_mappings(unwrapped_response)
    if not mappings:
        mappings = _extract_legacy_format_mappings(unwrapped_response)
    if not mappings:
        mappings = _extract_plain_key_value_mappings(unwrapped_response)
    return [mapping for mapping in mappings if isinstance(mapping, dict)]


def parse_ai_response(response_text: str) -> str:
    """
    Parse AI response text (JSON-first, fallback to legacy format).
    Returns JSON string with parsed mappings and metadata.
    """
    unwrapped_response = _extract_response_text(response_text)

    if not unwrapped_response or not unwrapped_response.strip():
        return json.dumps({"success": False, "error": "Empty response text", "mappings": []}, ensure_ascii=False)

    mappings = _extract_raw_mappings(unwrapped_response)

    # Validate extracted mappings.
    valid_mappings = []
    for mapping in mappings:
        if not isinstance(mapping, dict):
            continue

        normalized_mapping = _normalize_mapping_fields(mapping)

        # Check for required fields (allow some flexibility).
        if any(field in normalized_mapping for field in ["source_wb_keyword", "source_ws"]):
            valid_mappings.append(normalized_mapping)

    result = {
        "success": len(valid_mappings) > 0,
        "mappings_extracted": len(valid_mappings),
        "mappings": valid_mappings,
        "raw_input_length": len(response_text),
    }

    return json.dumps(result, ensure_ascii=False)


def append_mappings_to_config(config_file_path: str, response_text: str) -> str:
    """
    Parse AI response and append extracted mappings to Excel config file.
    Returns JSON string with append result metadata.
    """
    # Parse the response.
    parse_result_str = parse_ai_response(response_text)
    parse_result = json.loads(parse_result_str)

    if not parse_result["success"]:
        return json.dumps({
            "success": False,
            "error": "Failed to parse AI response",
            "rows_appended": 0,
            "by_sheet": {},
        }, ensure_ascii=False)

    mappings = parse_result["mappings"]
    config_path = _resolve_config_path(config_file_path)

    # Validate config file exists.
    if not config_path.exists():
        return json.dumps({
            "success": False,
            "error": f"Config file not found: {config_path}",
            "rows_appended": 0,
            "by_sheet": {},
        }, ensure_ascii=False)

    if config_path.suffix.lower() not in {".xlsx", ".xls", ".xlsm"}:
        return json.dumps({
            "success": False,
            "error": f"Unsupported file type: {config_path.suffix}",
            "rows_appended": 0,
            "by_sheet": {},
        }, ensure_ascii=False)

    try:
        wb = load_workbook(config_path)
    except Exception as e:
        return json.dumps({
            "success": False,
            "error": f"Failed to load workbook: {str(e)}",
            "rows_appended": 0,
            "by_sheet": {},
        }, ensure_ascii=False)

    rows_appended = 0
    by_sheet = {}
    target_sheet_name = _pick_target_sheet_name(wb)

    if not target_sheet_name:
        return json.dumps({
            "success": False,
            "error": "No worksheet found in config workbook.",
            "rows_appended": 0,
            "by_sheet": {},
        }, ensure_ascii=False)

    ws = wb[target_sheet_name]

    for mapping in mappings:
        source_ws = mapping.get("source_ws")
        row_data = [
            _safe_cell_value(mapping.get("source_wb_keyword")),
            _safe_cell_value(mapping.get("source_ws")),
            _safe_cell_value(mapping.get("item_col_position")),
            _safe_cell_value(mapping.get("item_col_name")),
            _safe_cell_value(mapping.get("forecast_col_position")),
            _safe_cell_value(mapping.get("forecast_col_name")),
            _safe_cell_value(mapping.get("data_start_row")),
        ]

        ws.append(row_data)
        rows_appended += 1
        by_sheet[target_sheet_name] = by_sheet.get(target_sheet_name, 0) + 1

    # Save the workbook.
    try:
        wb.save(config_path)
    except Exception as e:
        return json.dumps({
            "success": False,
            "error": f"Failed to save workbook: {str(e)}",
            "rows_appended": rows_appended,
            "by_sheet": by_sheet,
        }, ensure_ascii=False)

    return json.dumps({
        "success": True,
        "config_file": str(config_path),
        "rows_appended": rows_appended,
        "by_sheet": by_sheet,
    }, ensure_ascii=False)


def select_file(filetypes=None, title="Select file"):
    """
    Open file dialog and return selected file path.
    """
    if filetypes is None:
        filetypes = [
            ("JSON files", ("*.json",)),
            ("Excel files", ("*.xlsx", "*.xls", "*.xlsm")),
            ("Text files", ("*.txt",)),
            ("All files", ("*.*",)),
        ]

    # Import tkinter lazily so non-UI environments can still import this module.
    import tkinter as tk
    from tkinter import filedialog

    root = tk.Tk()
    root.withdraw()
    file_path = filedialog.askopenfilename(title=title, filetypes=filetypes)
    root.destroy()
    return file_path


if __name__ == "__main__":
    print("=" * 60)
    print ("select AI response file and config file to append mappings from response to config")

    # Get response file and config file from command line or prompt user.
    response_file = None
    config_file = None

    if len(sys.argv) > 1:
        response_file = sys.argv[1]
        print(f"Using response file from command line: {response_file}")
    else:
        print("Select the AI response file")
        response_file = select_file(
            filetypes=[
                ("JSON files", ("*.json",)),
                ("Text files", ("*.txt",)),
                ("All files", ("*.*",)),
            ],
            title="Select AI response file (JSON or text)"
        )

    if not response_file:
        print("No response file selected. Exiting.")
        exit(1)

    if len(sys.argv) > 2:
        config_file = sys.argv[2]
        print(f"Using config file from command line: {config_file}")
    else:
        print("Select the Excel config file")
        config_file = select_file(
            filetypes=[
                ("Excel files", ("*.xlsx", "*.xls", "*.xlsm")),
                ("All files", ("*.*",)),
            ],
            title="Select Excel config file"
        )

    if not config_file:
        print("No config file selected. Exiting.")
        exit(1)

    try:
        # Read response file.
        response_path = Path(response_file)
        if not response_path.exists():
            raise FileNotFoundError(f"Response file not found: {response_file}")

        with response_path.open() as f:
            response_text = f.read()

        print(f"Appending mappings to {config_file}...")
        result_str = append_mappings_to_config(config_file, response_text)
        result = json.loads(result_str)

        print("Result:")
        print(json.dumps(result, indent=2, ensure_ascii=False))

        if result["success"]:
            print(f"Successfully appended {result['rows_appended']} rows.")
        else:
            print(f"Error: {result.get('error', 'Unknown error')}")
            exit(1)

    except Exception as error:
        print(f"Error: {error}")
        exit(1)

    print("=" * 60)
